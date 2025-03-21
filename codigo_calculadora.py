import streamlit as st
import pandas as pd
import os
import re
from tabulate import tabulate
from fuzzywuzzy import process
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --------------------------
# 📁 Rutas
# --------------------------
ruta = r"C:\Users\Paula Venegas\OneDrive - IN TRUCKS SAS\Documentos - PROCESOS\6. Calculadora Financiera\CALCULADORA_FINANCIERA_PYTHON.xlsx"
ruta_salida = r"C:\Users\Paula Venegas\OneDrive - IN TRUCKS SAS\Documentos - PROCESOS\6. Calculadora Financiera\RESULTADO_CALCULADORA.xlsx"
ruta_mga_csv = r"C:\Users\Paula Venegas\OneDrive - IN TRUCKS SAS\Documentos - PROCESOS\6. Calculadora Financiera\MGA_APP_INFO.csv"
Ruta_nueva = ruta

#CREAR MENU DESPLEGABLES EN STREAMLIT
hoja = "Selección"

@st.cache_data
def cargar_datos():
    df = pd.read_excel(ruta, sheet_name=hoja)
    coberturas = df['Coverages'].dropna().unique()
    mgas = df['MGA FILTRADA'].dropna().unique()
    carriers = df["Carrier mayus"].dropna().unique()
    estados = df['Abreviación'].dropna().unique()
    return coberturas, mgas, carriers, estados



# Cargar datos
coberturas_list, mgas_list, carriers_list, estados_list = cargar_datos()

st.title("Calculadora Financiera - Ingreso de Coberturas")

# Inicializar número de filas
if "num_filas" not in st.session_state:
    st.session_state.num_filas = 1

# Botón para agregar otra cobertura
if st.button("➕ Adicionar otra cobertura"):
    st.session_state.num_filas += 1

# Guardar entradas
datos_coberturas = []

# Mostrar formularios por cobertura
st.subheader("Coberturas Ingresadas")

for i in range(st.session_state.num_filas):
    with st.expander(f"Cobertura #{i+1}", expanded=True):
        col1, col2 = st.columns(2)
        cobertura = col1.selectbox("Cobertura", sorted(coberturas_list), key=f"cobertura_{i}")
        carrier = col2.selectbox("Carrier", sorted(carriers_list), key=f"carrier_{i}")
        
        col3, col4 = st.columns(2)
        mga = col3.selectbox("MGA", sorted(mgas_list), key=f"mga_{i}")
        total_premium = col4.number_input("Total Premium", min_value=0.0, format="%.2f", key=f"premium_{i}")

        estado = st.selectbox("Estado", sorted(estados_list), key=f"estado_{i}")

        # Guardar datos de esta cobertura
        datos_coberturas.append({
            "Cobertura": cobertura,
            "Carrier": carrier,
            "MGA": mga,
            "Total Premium": total_premium,
            "Estado": estado
        })

# Mostrar resultados si hay datos
if datos_coberturas:
    st.markdown("### Vista previa de tus coberturas")
    df_preview = pd.DataFrame(datos_coberturas)
    st.dataframe(df_preview)




# --------------------------
# 🔧 Funciones auxiliares
# --------------------------
def limpiar_elemento(e):
    return e.strip().upper()

def convertir_a_lista(cobertura):
    if pd.isna(cobertura) or cobertura == '':
        return []
    return re.split(r'[\s,\-\+]+', str(cobertura).strip())

def estandarizar_lista(lista):
    lista_limpia = [limpiar_elemento(x) for x in lista if x]
    return ', '.join(sorted(lista_limpia))

def limpiar_texto(texto):
    if pd.isna(texto):
        return ""
    return str(texto).strip().lower().replace('\n', '').replace('\r', '')

# --------------------------
# 🚀 Inicio del proceso
# --------------------------
try:
    if not os.access(ruta, os.R_OK):
        raise PermissionError("❌ Archivo está en uso o no tienes permisos para acceder")

    # --------------------------
    # 📄 Leer Input_tabla
    # --------------------------
    df = pd.read_excel(ruta, sheet_name='Input_tabla', usecols='A:F', nrows=20)
    df.columns = df.columns.str.strip()
    print("✅ Archivo 'Input_tabla' leído correctamente")

    # Agrupar y estandarizar
    df_grouped = (
        df.groupby(['Carrier', 'MGA'], as_index=False)
        .agg({'Cobertura': list, 'Total Premium': 'sum', 'Estado': 'first'})
    )
    df_grouped['Cobertura'] = df_grouped['Cobertura'].apply(lambda x: estandarizar_lista(x))
    df_grouped.insert(0, 'N°', range(1, len(df_grouped) + 1))
    df_grouped.to_excel(ruta_salida, index=False)
    print(f"✅ Archivo guardado en: {ruta_salida}")

    # --------------------------
    # 📄 Leer Base_2024
    # --------------------------
    base_datos_ventas = pd.read_excel(ruta, sheet_name='Base_2024', usecols='A:BR')
    base_datos_ventas['Coverages'] = base_datos_ventas['Coverages'].fillna('').str.strip()
    base_datos_ventas = base_datos_ventas[base_datos_ventas['Coverages'].apply(lambda x: len(x) > 0)]
    base_datos_ventas['Coverages'] = base_datos_ventas['Coverages'].apply(lambda x: estandarizar_lista(convertir_a_lista(x)))
    base_datos_ventas['MGA - GA'] = base_datos_ventas['MGA - GA'].fillna('').str.strip().str.upper()
    base_datos_ventas['Carrier'] = base_datos_ventas['Carrier'].fillna('').str.strip().str.upper()
    base_datos_ventas['Total Premium'] = base_datos_ventas['Total Premium'].astype(str).str.replace('[\$,]', '', regex=True).str.strip()
    base_datos_ventas['Total Premium'] = pd.to_numeric(base_datos_ventas['Total Premium'], errors='coerce')
    df_grouped['MGA'] = df_grouped['MGA'].fillna('').str.strip().str.upper()
    df_grouped['Carrier'] = df_grouped['Carrier'].fillna('').str.strip().str.upper()
    df_grouped['Total Premium'] = pd.to_numeric(df_grouped['Total Premium'], errors='coerce')

    resultados_filtrados = []

    # --------------------------
    # 🔍 Filtrado inteligente
    # --------------------------
    for index, fila in df_grouped.iterrows():
        cobertura_filtrar = fila['Cobertura']
        mga_filtrar = fila['MGA']
        total_premium = fila['Total Premium']
        carrier_filtrar = fila['Carrier']
        estado_filtrar = fila['Estado']

        print(f"\n🔎 Fila {index + 1}: Estado={estado_filtrar}, Cobertura={cobertura_filtrar}, MGA={mga_filtrar}, Total Premium={total_premium}, Carrier={carrier_filtrar}")

        filtrado = base_datos_ventas[
            (base_datos_ventas['Coverages'] == cobertura_filtrar) &
            (base_datos_ventas['MGA - GA'] == mga_filtrar) &
            (base_datos_ventas['Carrier'] == carrier_filtrar) &
            (base_datos_ventas['Total Premium'] >= total_premium) &
            (base_datos_ventas['State'] == estado_filtrar)
        ]

        if not filtrado.empty:
            fila_min_premium = filtrado.loc[filtrado['Total Premium'].idxmin()]
            resultados_filtrados.append(fila_min_premium)
            print("✅ Coincidencia encontrada.")
        else:
            resultados_filtrados.append(None)
            print("❌ Sin coincidencias.")

    # --------------------------
    # 🧾 Enriquecer con MGA_APP_INFO
    # --------------------------
    try:
        df_mga_info = pd.read_csv(ruta_mga_csv)
        print("✅ CSV MGA_APP_INFO cargado.")
        df_mga_info['MGA'] = df_mga_info['MGA'].fillna('').str.strip().str.upper()
        df_mga_info['DOWN PAYMENT %'] = df_mga_info['DOWN PAYMENT %'].fillna('')
        df_mga_info['FINANCIAMIENTO'] = df_mga_info['FINANCIAMIENTO'].fillna('')
        mga_nombres_csv = df_mga_info['MGA'].dropna().str.upper().tolist()

        for i in range(len(resultados_filtrados)):
            if resultados_filtrados[i] is not None:
                fila_resultado = resultados_filtrados[i].copy()
                mga_nombre = fila_resultado.get('MGA - GA', '').strip().upper()
                mejor_coincidencia, score = process.extractOne(mga_nombre, mga_nombres_csv)

                if score >= 80:
                    info_mga = df_mga_info[df_mga_info['MGA'].str.upper() == mejor_coincidencia]
                    fila_resultado['Financiera'] = info_mga.iloc[0].get('FINANCIAMIENTO', '')
                    fila_resultado['% Downpayment'] = info_mga.iloc[0].get('DOWN PAYMENT %', '')
                else:
                    fila_resultado['Financiera'] = ''
                    fila_resultado['% Downpayment'] = ''
                resultados_filtrados[i] = fila_resultado
    except Exception as e:
        print(f"⚠️ Error leyendo MGA_APP_INFO: {e}")

    # --------------------------
    # 📝 Escribir en hoja "Calculadora"
    # --------------------------
    wb = load_workbook(Ruta_nueva)
    ws = wb['Calculadora']
    max_row = ws.max_row

    for row in range(2, max_row + 1):
        for col in range(1, 8):
            ws.cell(row=row, column=col).value = ""

    center_align = Alignment(horizontal="center")
    start_row = 2

    for resultado in resultados_filtrados:
        if resultado is not None:
            if isinstance(resultado, pd.Series):
                resultado = resultado.to_frame().T.iloc[0]

            ws[f'A{start_row}'] = resultado.get('Coverages', '')
            ws[f'B{start_row}'] = pd.to_numeric(resultado.get('Premium', 0), errors='coerce') or 0
            ws[f'C{start_row}'] = pd.to_numeric(resultado.get('Policy Fees', 0), errors='coerce') or 0
            ws[f'D{start_row}'] = pd.to_numeric(resultado.get('Policy Taxes', 0), errors='coerce') or 0
            ws[f'E{start_row}'] = resultado.get('MGA - GA', '')
            ws[f'F{start_row}'] = resultado.get('Financiera', '')
            ws[f'G{start_row}'] = resultado.get('% Downpayment', '')

            for col in 'ABCDEFG':
                ws[f'{col}{start_row}'].alignment = center_align

            start_row += 1

    # --------------------------
    # 📈 APR máximo en L2
    # --------------------------
    df_seleccion = pd.read_excel(Ruta_nueva, sheet_name='Selección', usecols='AC:AD')
    df_seleccion = df_seleccion.dropna(subset=['Carriers'])
    df_seleccion['Carriers_clean'] = df_seleccion['Carriers'].apply(limpiar_texto)
    df_seleccion['APR'] = df_seleccion['APR'].astype(str).str.replace('%', '', regex=False).astype(float)
    df_grouped['Carrier_clean'] = df_grouped['Carrier'].apply(limpiar_texto)

    df_comparado = df_grouped.merge(
        df_seleccion[['Carriers_clean', 'APR']],
        left_on='Carrier_clean',
        right_on='Carriers_clean',
        how='left'
    )

    apr_validos = df_comparado['APR'].dropna()
    if not apr_validos.empty:
        apr_maximo = apr_validos.max()
        ws['L2'] = apr_maximo
        ws['L2'].alignment = Alignment(horizontal="center")
        print(f"✅ APR máximo escrito en celda L2: {apr_maximo}%")
    else:
        print("❌ No se encontró APR para escribir.")

    # Guardar todo
    wb.save(Ruta_nueva)
    print("✅ Hoja 'Calculadora' actualizada y archivo guardado.")

except PermissionError as pe:
    print(f"⚠️ {pe}")

except Exception as e:
    print(f"❌ Error inesperado: {e}")